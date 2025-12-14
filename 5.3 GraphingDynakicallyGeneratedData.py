from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

# -------------------------------------------------
# insertData(filePath: str, dataString: str) -> bool
# Appends a line of data to a CSV file
# -------------------------------------------------
def insertData(filePath, dataString):
    try:
        with open(filePath, "a") as file:
            file.write(dataString + "\n")
        return True
    except Exception as e:
        print("Error writing to file:", e)
        return False


# -------------------------------------------------
# viewData(filePath: str) -> None
# Displays CSV file contents
# -------------------------------------------------
def viewData(filePath):
    try:
        print(f"You selected 2 at {datetime.now()}")
        print(f"The file {filePath}")
        with open(filePath, "r") as file:
            print(file.read())
    except Exception as e:
        print("Error reading file:", e)


# -------------------------------------------------
# convertData(value: float, spreadsheetType: str) -> float
# Converts input data to metric values
# -------------------------------------------------
def convertData(value, spreadsheetType):
    if spreadsheetType == "Temperature":
        return (value - 32) * 5 / 9
    elif spreadsheetType == "Weight":
        return value / 2.205
    elif spreadsheetType == "Rain":
        return value * 2.54
    else:
        return None


# -------------------------------------------------
# getInput() -> None
# Collects user data and stores in ZooData.csv
# -------------------------------------------------
def getInput():
    spreadsheetType = input("Enter spreadsheet type (Temperature, Weight, Rain): ").strip()
    numEntries = int(input("How many entries are you inputting?\n "))

    for _ in range(numEntries):
        dateVal = input("Enter a date:\n ")

        prompts = {
            "Temperature": "Enter the highest temp for the inputted date:\n",
            "Weight": "Enter the weight in pounds for the inputted date:\n",
            "Rain": "Enter the rain amount in inches for the inputted date:\n"
        }

        value = float(input(prompts.get(spreadsheetType)))
        converted = convertData(value, spreadsheetType)

        if converted is None:
            print("Invalid spreadsheet type. Entry skipped.")
            continue

        dataLine = f"{dateVal},{value},{converted}"

        if insertData("ZooData.csv", dataLine):
            print(f"The following data was saved at {datetime.now()}: {dataLine}")


# -------------------------------------------------
# createChart(filePath: str, chartType: str) -> None
# Writes CSV data to Excel and creates chart
# -------------------------------------------------
def createChart(filePath, chartType):

    dataChoice = input(
        "Choose data source:\n"
        "1 Initial Data\n"
        "2 Converted Data\n"
        "Enter 1 or 2: "
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Zoo Data"
    ws.append(["Date", "Value"])

    try:
        with open(filePath, "r") as file:
            for line in file:
                dateVal, original, converted = line.strip().split(",")

                if dataChoice == "1":
                    value = float(original)
                else:
                    if converted == "None":
                        continue
                    value = float(converted)

                ws.append([dateVal, value])

    except Exception as e:
        print("Error reading CSV:", e)
        return

    chart = BarChart() if chartType == "bar" else LineChart()

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.x_axis.title = "Date"
    chart.y_axis.title = "Value"
    chart.title = f"EriBek9068 {datetime.now().strftime('%m/%d/%Y')}"

    ws.add_chart(chart, "E2")
    wb.save("final.xlsx")

    print("Excel report created: final.xlsx")


# -------------------------------------------------
# generateReport(filePath: str) -> None
# Prompts user and generates Excel chart
# -------------------------------------------------
def generateReport(filePath):

    chartChoice = input("Choose chart type (bar or line): ").lower()

    if chartChoice not in ["bar", "line"]:
        print("Invalid chart type.")
        return

    createChart(filePath, chartChoice)


# -------------------------------------------------
# Main Menu
# -------------------------------------------------
print("EriBek9068's Spreadsheet Automation Menu")
menuOptions = ["Input Data", "View Current Data", "Generate Report"]

print("Choose a number from the following options:")
for index, option in enumerate(menuOptions, start=1):
    print(f"{index} {option}")

userChoice = input()

if userChoice == "1":
    getInput()
elif userChoice == "2":
    viewData("ZooData.csv")
elif userChoice == "3":
    generateReport("ZooData.csv")
else:
    print("Error: The chosen functionality is not implemented yet")
